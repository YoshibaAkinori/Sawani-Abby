#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
予約情報をExcelに追記するスクリプト（本番環境用）
"""

import sys
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def get_project_paths():
    """プロジェクトのパスを取得"""
    # スクリプトの場所から2階層上がプロジェクトルート
    project_root = Path(__file__).parent.parent
    
    # Excelファイル保存先
    base_dir = project_root / "data" / "excel"
    # ダウンロード用コピー先
    output_dir = project_root / "public" / "downloads"
    
    # ディレクトリが存在しない場合は作成
    base_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    return base_dir, output_dir

def create_excel_template(year, month):
    """Excelテンプレートを作成"""
    wb = Workbook()
    ws = wb.active
    ws.title = "総合売上"
    
    # スタイル定義
    header_font = Font(name='游ゴシック', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='游ゴシック', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 列幅設定
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['H'].width = 10
    
    # ヘッダー設定（13行目）
    headers = {
        'A13': '予約日',
        'C13': '予約者名',
        'E13': '担当者',
        'H13': '来店回数'
    }
    
    for cell, value in headers.items():
        ws[cell] = value
        ws[cell].font = header_font
        ws[cell].fill = header_fill
        ws[cell].alignment = header_alignment
        ws[cell].border = border
    
    return wb

def format_date(date_str):
    """日付を 'MM月DD日' 形式に変換"""
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return f"{date_obj.month}月{date_obj.day}日"
    except:
        return date_str

def update_excel(booking_data):
    """Excelファイルを更新"""
    try:
        # パス取得
        base_dir, output_dir = get_project_paths()
        
        # 予約日から年月を取得
        date_obj = datetime.strptime(booking_data['date'], '%Y-%m-%d')
        year = date_obj.year
        month = date_obj.month
        
        # ファイル名生成
        file_name = f"{year}年 {month:02d}月売上.xlsx"
        file_path = base_dir / file_name
        
        # ファイルが存在しない場合は新規作成
        if not file_path.exists():
            wb = create_excel_template(year, month)
            ws = wb.active
            start_row = 14
        else:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 次の空行を探す（14行目から）
            start_row = 14
            while ws[f'A{start_row}'].value is not None:
                start_row += 1
        
        # データを書き込み
        ws[f'A{start_row}'] = format_date(booking_data['date'])
        ws[f'C{start_row}'] = booking_data['customer_name']
        ws[f'E{start_row}'] = booking_data['staff_name']
        ws[f'H{start_row}'] = booking_data['visit_count']
        
        # スタイル適用
        data_font = Font(name='游ゴシック', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        for col in ['A', 'C', 'E', 'H']:
            cell = ws[f'{col}{start_row}']
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
        
        # 保存
        wb.save(file_path)
        
        # ダウンロード用にコピー
        output_path = output_dir / file_name
        wb.save(output_path)
        
        # 成功レスポンス
        result = {
            "success": True,
            "file_name": file_name,
            "row": start_row,
            "file_path": str(file_path),
            "output_path": str(output_path)
        }
        
        print(json.dumps(result, ensure_ascii=False))
        return 0
        
    except Exception as e:
        # エラーレスポンス
        error_result = {
            "success": False,
            "error": str(e)
        }
        print(json.dumps(error_result, ensure_ascii=False))
        return 1

def main():
    if len(sys.argv) < 2:
        error_result = {
            "success": False,
            "error": "引数が不足しています"
        }
        print(json.dumps(error_result, ensure_ascii=False))
        return 1
    
    try:
        booking_data = json.loads(sys.argv[1])
        return update_excel(booking_data)
    except json.JSONDecodeError as e:
        error_result = {
            "success": False,
            "error": f"JSON解析エラー: {str(e)}"
        }
        print(json.dumps(error_result, ensure_ascii=False))
        return 1

if __name__ == "__main__":
    sys.exit(main())